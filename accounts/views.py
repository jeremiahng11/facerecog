import json
import logging
import mimetypes
import os
import tempfile
from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.files import File as DjangoFile
from django.core.paginator import Paginator
from django.http import JsonResponse, Http404, FileResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.conf import settings
from django.db.models import Count, Q

from .models import StaffUser, FaceLoginLog, AdminActionLog, KioskConfig
from .forms import StaffLoginForm, StaffUserCreationForm, StaffUserEditForm, FacePhotoUploadForm
from . import face_utils

logger = logging.getLogger(__name__)


def is_admin(user):
    return user.is_staff or user.is_superuser


def parse_device(request) -> str:
    """Parse User-Agent into a human-readable device string."""
    ua = request.META.get('HTTP_USER_AGENT', '')
    if not ua:
        return 'Unknown'
    if 'iPhone' in ua:
        platform = 'iOS Mobile'
    elif 'iPad' in ua:
        platform = 'iOS Tablet'
    elif 'Android' in ua:
        platform = 'Android Mobile' if 'Mobile' in ua else 'Android Tablet'
    elif 'Windows' in ua:
        platform = 'Windows'
    elif 'Macintosh' in ua or 'Mac OS' in ua:
        platform = 'macOS'
    elif 'Linux' in ua:
        platform = 'Linux'
    elif 'CrOS' in ua:
        platform = 'Chrome OS'
    else:
        platform = 'Unknown OS'
    if 'Edg/' in ua or 'Edge/' in ua:
        browser = 'Edge'
    elif 'OPR/' in ua or 'Opera' in ua:
        browser = 'Opera'
    elif 'Firefox/' in ua:
        browser = 'Firefox'
    elif 'CriOS/' in ua or ('Chrome/' in ua and 'Safari/' in ua):
        browser = 'Chrome'
    elif 'Safari/' in ua and 'Chrome/' not in ua:
        browser = 'Safari'
    else:
        browser = 'Unknown Browser'
    return f'{platform} · {browser}'


def get_client_ip(request):
    """Return real client IP from X-Forwarded-For or REMOTE_ADDR."""
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


# ─── Rate Limiting & IP Lockout ──────────────────────────────────────────────
# Simple in-DB rate limiter. No extra dependencies needed.

def _is_ip_locked_out(ip_addr):
    """Check if an IP is temporarily locked out due to too many failed scan sessions."""
    threshold = getattr(settings, 'FACE_LOCKOUT_THRESHOLD', 15)
    lockout_minutes = getattr(settings, 'FACE_LOCKOUT_DURATION_MINUTES', 5)
    cutoff = timezone.now() - timedelta(minutes=lockout_minutes)

    recent_fails = FaceLoginLog.objects.filter(
        ip_address=ip_addr,
        success=False,
        timestamp__gte=cutoff,
    ).count()

    return recent_fails >= threshold


def _send_security_notification(subject, message_body):
    """Send email notification for security events (non-blocking)."""
    admin_email = getattr(settings, 'ADMIN_NOTIFICATION_EMAIL', '')
    if not admin_email:
        return
    try:
        from django.core.mail import send_mail
        send_mail(
            subject=f'[FaceID Portal] {subject}',
            message=message_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[admin_email],
            fail_silently=True,
        )
    except Exception as e:
        logger.warning(f"Failed to send security notification: {e}")


def _log_admin_action(admin_user, action, target_user, details=''):
    """Create an admin audit log entry."""
    AdminActionLog.objects.create(
        admin_user=admin_user,
        action=action,
        target_staff_id=target_user.staff_id if hasattr(target_user, 'staff_id') else str(target_user),
        target_name=target_user.full_name if hasattr(target_user, 'full_name') else '',
        details=details,
    )


# ─── Login / Logout ────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')  # smart router dispatches by role
    if request.method == 'POST':
        form = StaffLoginForm(request.POST)
        if form.is_valid():
            staff_id = form.cleaned_data['staff_id']
            password = form.cleaned_data['password']
            user = authenticate(request, username=staff_id, password=password)
            if user is not None and user.is_active:
                # Auto-disable expired temp/intern accounts at login time
                if (user.staff_type in ('temp', 'intern')
                        and user.contract_end_date
                        and user.contract_end_date < timezone.localdate()):
                    user.is_active = False
                    user.save(update_fields=['is_active'])
                    messages.error(request, 'Your account has expired. Please contact admin.')
                else:
                    login(request, user)
                    return redirect('dashboard')
            else:
                messages.error(request, 'Invalid Staff ID or Password.')
    else:
        form = StaffLoginForm()
    return render(request, 'accounts/login.html', {'form': form})


def logout_view(request):
    logout(request)
    next_url = request.POST.get('next') or request.GET.get('next') or ''
    # Only allow safe, relative, in-app redirects.
    if next_url.startswith('/') and not next_url.startswith('//'):
        return redirect(next_url)
    return redirect('login')


# ─── Face ID Login ─────────────────────────────────────────────────────────────

def face_login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'accounts/face_login.html')


@require_POST
def face_verify_ajax(request):
    """
    AJAX: receive 2 base64 frames in ONE request, compare against enrolled
    faces, run liveness check, and grant login — all in a single round-trip.

    The client captures two frames ~300ms apart and sends both together,
    eliminating the "Verifying 1/2 → 2/2" multi-request flow that caused
    perceived slowness (each round-trip was 1-3s on shared Railway CPU).
    """
    try:
        ip_addr = get_client_ip(request)
        device_info = parse_device(request)

        if _is_ip_locked_out(ip_addr):
            return JsonResponse({
                'success': False,
                'message': 'Too many failed attempts. Please try again later.',
                'locked_out': True,
            })

        data = json.loads(request.body)
        # Accept both single-frame (image) and dual-frame (images) payloads
        images = data.get('images') or []
        if not images and data.get('image'):
            images = [data['image']]
        if not images:
            return JsonResponse({'success': False, 'message': 'No image received'})

        min_confidence = getattr(settings, 'FACE_MIN_CONFIDENCE', 65)
        tolerance = getattr(settings, 'FACE_RECOGNITION_TOLERANCE', 0.4)

        # Extract encodings. Frame1 runs full HOG detection + quality checks;
        # subsequent frames reuse frame1's face location to skip re-detection.
        encodings = []
        frame1_location = None
        for i, img in enumerate(images):
            if i == 0:
                result = face_utils.validate_and_extract(img)
                if not result['ok']:
                    return JsonResponse({
                        'success': False,
                        'message': result['reason'],
                        'face_detected': False,
                    })
                encodings.append(result['encoding'])
                frame1_location = result.get('location')
            else:
                enc = face_utils.fast_extract(img, known_location=frame1_location)
                if enc is None:
                    return JsonResponse({
                        'success': False,
                        'message': 'Hold still…',
                        'face_detected': False,
                    })
                encodings.append(enc)

        # ── Batch compare first frame against all enrolled users ──
        match = face_utils.encoding_cache.find_best_match(encodings[0], tolerance)

        if not match or match['confidence'] < min_confidence:
            return JsonResponse({
                'success': False, 'face_detected': True,
                'message': 'Face not recognised. Try again or use Staff ID login.',
                'confidence': 0,
            })

        # If only 1 frame was sent (legacy client), ask for more
        if len(encodings) < 2:
            return JsonResponse({
                'success': False,
                'face_detected': True,
                'message': 'Verifying…',
                'confidence': match['confidence'],
                'verifying': True,
            })

        # ── Verify second frame matches the same user ─────────────
        match2 = face_utils.encoding_cache.find_best_match(encodings[1], tolerance)
        if not match2 or match2['staff_id'] != match['staff_id']:
            return JsonResponse({
                'success': False, 'face_detected': True,
                'message': 'Hold still — verifying…',
                'verifying': True,
            })

        # ── Liveness check: verify frame variance ─────────────────
        if not face_utils.check_encoding_variance(encodings, min_std=0.008):
            FaceLoginLog.objects.create(
                success=False, ip_address=ip_addr,
                device=device_info,
                notes='Liveness check failed — possible photo spoofing',
            )
            logger.warning(f"Liveness check failed during login from {ip_addr}")
            _send_security_notification(
                'Possible Photo Spoofing Attempt',
                f'A face login attempt from IP {ip_addr} failed the '
                f'liveness check (zero encoding variance across '
                f'{len(encodings)} frames). This may indicate a '
                f'printed photo or phone screen was used.\n'
                f'Device: {device_info}\n'
                f'Matched user: {match["staff_id"]}\n'
                f'Time: {timezone.now().isoformat()}',
            )
            return JsonResponse({
                'success': False,
                'face_detected': True,
                'message': 'Liveness check failed. Please look directly at the camera and blink naturally.',
            })

        # ── Grant login ───────────────────────────────────────────
        best_match = StaffUser.objects.get(pk=match['pk'])

        if not best_match.is_active:
            return JsonResponse({
                'success': False,
                'face_detected': True,
                'message': 'Your account has been deactivated. Please contact admin.',
            })

        # Auto-disable expired temp/intern accounts
        if (best_match.staff_type in ('temp', 'intern')
                and best_match.contract_end_date
                and best_match.contract_end_date < timezone.localdate()):
            best_match.is_active = False
            best_match.save(update_fields=['is_active'])
            return JsonResponse({
                'success': False,
                'face_detected': True,
                'message': 'Your account has expired. Please contact admin.',
            })

        FaceLoginLog.objects.create(
            user=best_match, success=True,
            confidence=match['confidence'], ip_address=ip_addr,
            device=device_info,
            notes='2-frame batch verify',
        )
        best_match.last_face_login = timezone.now()
        best_match.save(update_fields=['last_face_login'])
        login(request, best_match, backend='django.contrib.auth.backends.ModelBackend')

        logger.info(f"Face login: {best_match.staff_id} from {ip_addr} ({device_info})")

        return JsonResponse({
            'success': True,
            'message': f'Welcome, {best_match.display_name}!',
            'confidence': match['confidence'],
            'redirect': '/dashboard/',
        })

    except Exception as e:
        logger.exception("Error in face_verify_ajax")
        return JsonResponse({'success': False, 'message': f'Server error: {str(e)}'})


def _clear_face_session(request):
    """Remove temporary face verification state from the session."""
    for key in ('_face_match_user', '_face_match_count', '_face_match_encs'):
        request.session.pop(key, None)


@require_POST
def face_verify_fail_ajax(request):
    """
    AJAX: called once when a scan session exhausts its attempts without
    finding a match. Logs a single failure entry for the IP lockout system.
    Separated from face_verify_ajax so we don't log every individual frame.
    """
    ip_addr = get_client_ip(request)
    device_info = parse_device(request)

    FaceLoginLog.objects.create(
        success=False,
        ip_address=ip_addr,
        device=device_info,
        notes='Scan session timeout — no match found',
    )

    if _is_ip_locked_out(ip_addr):
        _send_security_notification(
            'IP Locked Out — Too Many Failed Face Login Attempts',
            f'IP {ip_addr} has been temporarily locked out after '
            f'exceeding {settings.FACE_LOCKOUT_THRESHOLD} failed '
            f'face scan sessions.\nDevice: {device_info}\n'
            f'Time: {timezone.now().isoformat()}',
        )

    return JsonResponse({'logged': True})


# ─── Dashboard ─────────────────────────────────────────────────────────────────

@login_required
def dashboard_view(request):
    """
    Smart router: after login, land each role on the right home screen.
    Regular staff go straight to the PWA; admin/kitchen roles go to their
    respective dashboards.
    """
    user = request.user
    if user.is_admin_role:
        return redirect('cafeteria_dashboard')
    if getattr(user, 'role', '') in ('kitchen', 'cafe_bar', 'kitchen_admin'):
        return redirect('cafeteria_displays_hub')
    # Regular staff → PWA home (My Orders)
    return redirect('staff_portal_home')


# ─── Profile / Face Enrolment ──────────────────────────────────────────────────

@login_required
def profile_view(request):
    user = request.user
    if request.method == 'POST':
        form = FacePhotoUploadForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated.')
            return redirect('profile')
    else:
        form = FacePhotoUploadForm(instance=user)
    context = {
        'form': form,
        'user': user,
        'enroll_num_samples': getattr(settings, 'FACE_ENROLL_NUM_SAMPLES', 5),
        'requirements': [
            'Good lighting',
            'Face clearly visible',
            'Look directly at camera',
            'Only one person in frame',
            'No heavy filters or masks',
        ],
    }
    return render(request, 'accounts/profile.html', context)


@login_required
def my_face_photo_view(request):
    """Serve the current user's own face photo (owner-only)."""
    user = request.user
    if not user.face_photo:
        raise Http404("No face photo on file")
    try:
        f = user.face_photo.open('rb')
        content_type = mimetypes.guess_type(user.face_photo.name)[0] or 'image/jpeg'
        response = FileResponse(f, content_type=content_type)
        response['Cache-Control'] = 'private, no-store'
        return response
    except Exception:
        raise Http404("Face photo not accessible")


@login_required
@require_POST
def enroll_face_ajax(request):
    """
    AJAX: multi-sample enrollment with quality checks, liveness
    validation (frame variance), duplicate face check, and jittered
    encoding averaging.
    """
    try:
        data = json.loads(request.body)
        images = data.get('images', [])
        if not images and data.get('image'):
            images = [data['image']]

        num_samples_required = getattr(settings, 'FACE_ENROLL_NUM_SAMPLES', 3)
        dup_tolerance = getattr(settings, 'FACE_ENROLL_DUPLICATE_TOLERANCE', 0.35)

        if len(images) < num_samples_required:
            return JsonResponse({
                'success': False,
                'message': f'Need {num_samples_required} captures, got {len(images)}.',
            })

        user = request.user
        encodings = []

        for idx, image_data in enumerate(images):
            # Single-pass: validate quality + extract encoding together.
            # This avoids calling face_locations() twice per frame.
            result = face_utils.validate_and_extract(image_data)
            if not result['ok']:
                return JsonResponse({
                    'success': False,
                    'message': f'Capture {idx + 1}: {result["reason"]}',
                })
            encodings.append(result['encoding'])

        # ── Liveness: check frame variance ────────────────────────
        # If all encodings are nearly identical, the user might be
        # holding a static photo. Real faces produce slight encoding
        # variation across frames due to micro-movements. We check
        # that the standard deviation across samples exceeds a
        # minimum threshold.
        if not face_utils.check_encoding_variance(encodings, min_std=0.01):
            return JsonResponse({
                'success': False,
                'message': (
                    'Liveness check failed. Please move naturally and '
                    'blink while capturing. Static images are not accepted.'
                ),
            })

        final_encoding = face_utils.average_encodings(encodings)

        # ── Duplicate face check ──────────────────────────────────
        dup = face_utils.check_duplicate_face(
            final_encoding, tolerance=dup_tolerance,
            exclude_user_pk=user.pk,
        )
        if dup:
            dup_user = dup['user']
            logger.warning(
                f"Enrollment rejected for {user.staff_id}: face too similar "
                f"to {dup_user.staff_id} (distance={dup['distance']})"
            )
            _send_security_notification(
                'Face Enrollment Rejected — Duplicate Face',
                f'User {user.staff_id} ({user.full_name}) attempted to '
                f'enroll but their face was too similar to '
                f'{dup_user.staff_id} ({dup_user.full_name}).\n'
                f'Distance: {dup["distance"]}\n'
                f'Time: {timezone.now().isoformat()}',
            )
            return JsonResponse({
                'success': False,
                'message': (
                    'Enrollment failed: your face is too similar to another '
                    'enrolled user. Please contact an administrator.'
                ),
            })

        # ── Save face photo ───────────────────────────────────────
        filename = f"{user.staff_id}_face.jpg"
        with tempfile.TemporaryDirectory() as tmp_dir:
            save_path = os.path.join(tmp_dir, filename)
            saved = face_utils.save_face_snapshot(images[0], save_path)
            if saved:
                with open(save_path, 'rb') as f:
                    user.face_photo.save(
                        f'face_photos/{filename}', DjangoFile(f), save=False
                    )

        user.set_face_encoding(final_encoding)
        user.face_registered = True
        user.face_enabled = True
        user.save(update_fields=[
            'face_photo', 'face_encoding', 'face_registered', 'face_enabled'
        ])
        face_utils.encoding_cache.invalidate()

        return JsonResponse({
            'success': True,
            'message': 'Face enrolled successfully! You can now use Face ID login.',
        })

    except Exception as e:
        logger.exception("Error in enroll_face_ajax")
        return JsonResponse({'success': False, 'message': str(e)})


# ─── Kiosk Mode ──────────────────────────────────────────────────────────────

def kiosk_view(request):
    """
    Full-screen face scanner for shared devices (entrance tablets, kiosks).
    Continuously scans → shows welcome greeting on match → auto-resets
    after a few seconds. No login/auth required to render the page — the
    face_verify_ajax endpoint handles the actual authentication.
    """
    return render(request, 'accounts/kiosk.html')


# ─── Password Reset ──────────────────────────────────────────────────────────

def password_reset_view(request):
    """
    Custom password reset view that wraps Django's built-in form but
    catches SMTP errors and shows them to the user instead of silently
    failing or returning a 500.
    """
    from django.contrib.auth.forms import PasswordResetForm

    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            try:
                form.save(
                    request=request,
                    use_https=request.is_secure(),
                    email_template_name='accounts/password_reset_email.html',
                    subject_template_name='accounts/password_reset_subject.txt',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                )
                return redirect('password_reset_done')
            except Exception as e:
                logger.exception("Password reset email failed")
                messages.error(
                    request,
                    f'Failed to send reset email. Please contact an administrator. '
                    f'(Error: {type(e).__name__})'
                )
    else:
        form = None

    return render(request, 'accounts/password_reset.html', {'form': form})


# ─── Admin: User Management ────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def admin_users_view(request):
    users = StaffUser.objects.all().order_by('-date_joined')
    # Hide root users from anyone who isn't themselves a root user.
    if not getattr(request.user, 'is_root', False):
        users = users.filter(is_root=False)
    return render(request, 'accounts/admin_users.html', {'users': users})


@login_required
@user_passes_test(is_admin)
def admin_add_user_view(request):
    from decimal import Decimal
    cfg = KioskConfig.get()
    working_days = cfg.credit_working_days

    if request.method == 'POST':
        form = StaffUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])

            # ── Credit assignment ─────────────────────────────────
            prorate = form.cleaned_data.get('prorate_credit', True)
            if prorate:
                # Prorate: remaining days in the month / working days
                today = timezone.localdate()
                import calendar
                _, days_in_month = calendar.monthrange(today.year, today.month)
                remaining_days = days_in_month - today.day + 1  # include today
                ratio = Decimal(str(remaining_days)) / Decimal(str(working_days))
                # Cap at 1.0 so credit never exceeds monthly amount
                if ratio > 1:
                    ratio = Decimal('1')
                user.credit_balance = (user.monthly_credit * ratio).quantize(Decimal('0.01'))
            else:
                manual = form.cleaned_data.get('manual_credit')
                if manual is not None:
                    user.credit_balance = manual
                else:
                    user.credit_balance = user.monthly_credit

            user.save()
            _log_admin_action(request.user, 'create', user,
                              details=f'Credit: S${user.credit_balance}')
            messages.success(
                request,
                f'User {user.staff_id} created with S${user.credit_balance} credit.',
            )
            return redirect('admin_users')
    else:
        form = StaffUserCreationForm()

    return render(request, 'accounts/admin_add_user.html', {
        'form': form,
        'working_days': working_days,
    })


@login_required
@user_passes_test(is_admin)
def admin_edit_user_view(request, user_id):
    target_user = get_object_or_404(StaffUser, pk=user_id)
    # Non-root admins cannot edit root users — pretend they don't exist.
    if target_user.is_root and not getattr(request.user, 'is_root', False):
        raise Http404()
    if request.method == 'POST':
        form = StaffUserEditForm(request.POST, request.FILES, instance=target_user)
        if form.is_valid():
            form.save()
            _log_admin_action(request.user, 'edit', target_user)
            messages.success(request, 'User updated successfully.')
            return redirect('admin_users')
    else:
        form = StaffUserEditForm(instance=target_user)
    return render(request, 'accounts/admin_edit_user.html', {
        'form': form,
        'target_user': target_user,
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_user_view(request, user_id):
    target_user = get_object_or_404(StaffUser, pk=user_id)
    # Non-root admins cannot see or delete root users.
    if target_user.is_root and not getattr(request.user, 'is_root', False):
        raise Http404()
    if request.method == 'POST':
        _log_admin_action(request.user, 'delete', target_user)
        name = target_user.staff_id
        target_user.delete()
        messages.success(request, f'User {name} deleted.')
        return redirect('admin_users')
    return render(request, 'accounts/admin_confirm_delete.html', {'target_user': target_user})


@login_required
@user_passes_test(is_admin)
def admin_bulk_import_view(request):
    """CSV/Excel bulk import of users."""
    import csv
    import io
    import calendar
    from decimal import Decimal

    results = None
    if request.method == 'POST' and request.FILES.get('import_file'):
        upload = request.FILES['import_file']
        fname = upload.name.lower()

        try:
            # ── Parse file into list of dicts ─────────────────────
            if fname.endswith('.xlsx') or fname.endswith('.xls'):
                rows = _parse_excel(upload)
            else:
                decoded = upload.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)

            results = {'created': [], 'skipped': [], 'errors': []}

            # Prorate credit for bulk-imported users
            working_days = KioskConfig.get().credit_working_days
            default_credit = Decimal(str(getattr(settings, 'DEFAULT_MONTHLY_CREDIT', 50)))
            today = timezone.localdate()
            _, days_in_month = calendar.monthrange(today.year, today.month)
            remaining = days_in_month - today.day + 1
            ratio = min(Decimal('1'), Decimal(str(remaining)) / Decimal(str(working_days)))
            prorated = (default_credit * ratio).quantize(Decimal('0.01'))

            for i, row in enumerate(rows, start=2):  # row 1 is header
                staff_id = (row.get('staff_id') or '').strip()
                email = (row.get('email') or '').strip()
                full_name = (row.get('full_name') or '').strip()
                department = (row.get('department') or '').strip()
                password = (row.get('password') or '').strip()
                staff_type = (row.get('staff_type') or '').strip().lower()
                end_date_str = (row.get('contract_end_date') or '').strip()

                if not staff_id or not email or not password:
                    results['errors'].append(f'Row {i}: missing staff_id, email, or password')
                    continue

                if StaffUser.objects.filter(staff_id=staff_id).exists():
                    results['skipped'].append(f'{staff_id} — already exists')
                    continue
                if StaffUser.objects.filter(email=email).exists():
                    results['skipped'].append(f'{email} — email already in use')
                    continue

                # Validate staff_type
                if staff_type not in ('permanent', 'temp', 'intern', ''):
                    results['errors'].append(f'Row {i} ({staff_id}): invalid staff_type "{staff_type}"')
                    continue

                # Parse contract end date
                contract_end = None
                if end_date_str:
                    from datetime import datetime as _dt
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                        try:
                            contract_end = _dt.strptime(end_date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    if contract_end is None:
                        results['errors'].append(f'Row {i} ({staff_id}): invalid date "{end_date_str}" — use YYYY-MM-DD')
                        continue

                try:
                    user = StaffUser.objects.create_user(
                        staff_id=staff_id,
                        email=email,
                        password=password,
                        full_name=full_name,
                        department=department,
                        monthly_credit=default_credit,
                        credit_balance=prorated,
                        staff_type=staff_type or 'permanent',
                        contract_end_date=contract_end,
                    )
                    _log_admin_action(request.user, 'create', user, f'Bulk import · S${prorated}')
                    results['created'].append(staff_id)
                except Exception as e:
                    results['errors'].append(f'Row {i} ({staff_id}): {e}')

        except Exception as e:
            results = {'created': [], 'skipped': [], 'errors': [f'File parse error: {e}']}

    return render(request, 'accounts/admin_bulk_import.html', {'results': results})


def _parse_excel(upload):
    """Parse an Excel (.xlsx) file into a list of dicts (same as csv.DictReader output)."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError('Excel support requires openpyxl. Install with: pip install openpyxl')

    wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or '').strip().lower() for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        d = {}
        for j, val in enumerate(row):
            if j < len(headers):
                d[headers[j]] = str(val) if val is not None else ''
        result.append(d)
    wb.close()
    return result


@login_required
@user_passes_test(is_admin)
def admin_bulk_template_download_view(request):
    """Download a CSV/Excel template for bulk user import."""
    import csv as _csv
    from django.http import HttpResponse

    fmt = request.GET.get('format', 'csv')

    if fmt == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Staff Import'
            headers = ['staff_id', 'email', 'full_name', 'department', 'password', 'staff_type', 'contract_end_date']
            ws.append(headers)
            ws.append(['EMP-001', 'john@company.com', 'John Smith', 'Engineering', 'SecurePass123', 'permanent', ''])
            ws.append(['TMP-001', 'jane@company.com', 'Jane Doe', 'Marketing', 'TempPass456', 'temp', '2026-06-30'])
            ws.append(['INT-001', 'bob@company.com', 'Bob Lee', 'Finance', 'InternPass789', 'intern', '2026-08-31'])
            # Column widths
            for col, w in zip('ABCDEFG', [12, 24, 18, 16, 16, 12, 18]):
                ws.column_dimensions[col].width = w

            from io import BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="staff_import_template.xlsx"'
            return response
        except ImportError:
            pass  # Fall through to CSV

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_import_template.csv"'
    writer = _csv.writer(response)
    writer.writerow(['staff_id', 'email', 'full_name', 'department', 'password', 'staff_type', 'contract_end_date'])
    writer.writerow(['EMP-001', 'john@company.com', 'John Smith', 'Engineering', 'SecurePass123', 'permanent', ''])
    writer.writerow(['TMP-001', 'jane@company.com', 'Jane Doe', 'Marketing', 'TempPass456', 'temp', '2026-06-30'])
    writer.writerow(['INT-001', 'bob@company.com', 'Bob Lee', 'Finance', 'InternPass789', 'intern', '2026-08-31'])
    return response


@login_required
@user_passes_test(is_admin)
def admin_face_logs_view(request):
    logs_qs = FaceLoginLog.objects.all().select_related('user')
    paginator = Paginator(logs_qs, 50)
    page_number = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_number)
    return render(request, 'accounts/admin_face_logs.html', {
        'logs': logs_page,
        'paginator': paginator,
    })


@login_required
@user_passes_test(is_admin)
def admin_action_logs_view(request):
    """View admin action audit trail."""
    logs_qs = AdminActionLog.objects.all().select_related('admin_user')
    paginator = Paginator(logs_qs, 50)
    page_number = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_number)
    return render(request, 'accounts/admin_action_logs.html', {
        'logs': logs_page,
        'paginator': paginator,
    })


@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    """Admin face login activity summary."""
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)

    today_logs = FaceLoginLog.objects.filter(timestamp__gte=today_start)
    week_logs = FaceLoginLog.objects.filter(timestamp__gte=week_start)

    context = {
        'total_today': today_logs.count(),
        'success_today': today_logs.filter(success=True).count(),
        'failed_today': today_logs.filter(success=False).count(),
        'total_week': week_logs.count(),
        'success_week': week_logs.filter(success=True).count(),
        'failed_week': week_logs.filter(success=False).count(),
        'total_users': StaffUser.objects.filter(is_active=True).count(),
        'enrolled_users': StaffUser.objects.filter(face_registered=True, is_active=True).count(),
        'top_devices': (
            week_logs.exclude(device='')
            .values('device')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
        ),
        'recent_failures': (
            FaceLoginLog.objects.filter(success=False)
            .select_related('user')
            .order_by('-timestamp')[:10]
        ),
    }
    return render(request, 'accounts/admin_dashboard.html', context)


# ─── Re-enroll face for a specific user (admin) ───────────────────────────────

@login_required
@user_passes_test(is_admin)
@require_POST
def admin_reencode_user(request, user_id):
    target_user = get_object_or_404(StaffUser, pk=user_id)
    if target_user.face_photo:
        encoding = face_utils.extract_encoding_from_field_file(target_user.face_photo)
        if encoding:
            target_user.set_face_encoding(encoding)
            target_user.face_registered = True
            target_user.save()
            face_utils.encoding_cache.invalidate()
            _log_admin_action(request.user, 'reencode', target_user)
            messages.success(request, f'Face re-encoded for {target_user.staff_id}.')
        else:
            messages.error(request, 'No face detected in existing photo.')
    else:
        messages.error(request, 'No face photo on file.')
    return redirect('admin_edit_user', user_id=user_id)
